"""Report generation utilities for the Data Comparison Tool."""
import pandas as pd
import numpy as np
import os
import zipfile
from pathlib import Path
from datetime import datetime
from typing import Dict, Any, List, Tuple
import logging
from openpyxl.styles import PatternFill, Font, Color
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

class ReportGenerator:
    def __init__(self, output_dir: str = "reports"):
        """Initialize the report generator."""
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
        
    def execute_query(self, df: pd.DataFrame, query: str = None) -> pd.DataFrame:
        """Execute a SQL-like query on the DataFrame.
        
        Args:
            df: The DataFrame to query
            query: SQL-like query string (e.g., "SELECT * FROM data WHERE Column > 5")
                  If None, returns the original DataFrame
        
        Returns:
            Filtered DataFrame based on the query
        """
        if not query:
            return df
            
        try:
            # Parse the query
            query = query.strip().lower()
            if not query.startswith("select"):
                raise ValueError("Query must start with SELECT")
                
            # Extract the WHERE clause if it exists
            where_clause = None
            if "where" in query:
                where_parts = query.split("where")
                if len(where_parts) != 2:
                    raise ValueError("Invalid WHERE clause")
                query = where_parts[0]
                where_clause = where_parts[1].strip()
            
            # Extract column names
            cols_part = query.replace("select", "").strip()
            if cols_part == "*":
                selected_cols = df.columns.tolist()
            else:
                selected_cols = [col.strip() for col in cols_part.split(",")]
                
            # Validate columns exist
            missing_cols = [col for col in selected_cols if col not in df.columns]
            if missing_cols:
                raise ValueError(f"Columns not found: {', '.join(missing_cols)}")
            
            result_df = df[selected_cols]
            
            # Apply WHERE clause if it exists
            if where_clause:
                # Replace column names with df[] notation
                for col in df.columns:
                    where_clause = where_clause.replace(col, f"df['{col}']")
                result_df = result_df[eval(where_clause)]
            
            return result_df
            
        except Exception as e:
            logger.error(f"Error executing query: {str(e)}")
            raise ValueError(f"Query execution failed: {str(e)}")
        
    def _style_excel_cell(self, cell, is_pass: bool):
        """Apply styling to Excel cell based on pass/fail status."""
        if is_pass:
            cell.fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')  # Light green
            cell.font = Font(color='006400')  # Dark green
        else:
            cell.fill = PatternFill(start_color='FFB6C1', end_color='FFB6C1', fill_type='solid')  # Light pink
            cell.font = Font(color='8B0000')  # Dark red

    def _calculate_aggregations(self, df: pd.DataFrame, numeric_cols: List[str]) -> pd.DataFrame:
        """Calculate aggregations for numeric columns."""
        aggs = []
        for col in numeric_cols:
            try:
                aggs.append({
                    'Column': col,
                    'Sum': df[col].sum(),
                    'Mean': df[col].mean(),
                    'Min': df[col].min(),
                    'Max': df[col].max(),
                    'StdDev': df[col].std()
                })
            except Exception as e:
                logger.error(f"Error calculating aggregations for column {col}: {str(e)}")
                # Add a placeholder with error indicators
                aggs.append({
                    'Column': col,
                    'Sum': np.nan,
                    'Mean': np.nan,
                    'Min': np.nan,
                    'Max': np.nan,
                    'StdDev': np.nan
                })
        return pd.DataFrame(aggs)

    def generate_regression_report(self, comparison_results: Dict[str, Any], 
                                 source_df: pd.DataFrame, 
                                 target_df: pd.DataFrame) -> str:
        """Generate enhanced regression report with multiple checks."""
        try:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            report_path = self.output_dir / f"regression_report_{timestamp}.xlsx"
            
            with pd.ExcelWriter(str(report_path), engine='openpyxl') as writer:
                # 1. Count Check Tab
                source_count = len(source_df)
                target_count = len(target_df)
                count_data = {
                    'Metric': ['Source Count', 'Target Count'],
                    'Value': [source_count, target_count],
                    'Result': ['Pass' if source_count == target_count else 'Fail',
                             'Pass' if source_count == target_count else 'Fail']
                }
                count_df = pd.DataFrame(count_data)
                count_df.to_excel(writer, sheet_name='CountCheck', index=False)
                
                # Apply conditional formatting
                count_sheet = writer.sheets['CountCheck']
                for idx, result in enumerate(count_df['Result'], start=2):
                    cell = count_sheet.cell(row=idx, column=3)
                    self._style_excel_cell(cell, result.lower() == 'pass')

                # 2. Aggregation Check Tab
                # Get numeric columns from both dataframes
                source_numeric = set(source_df.select_dtypes(include=[np.number]).columns)
                target_numeric = set(target_df.select_dtypes(include=[np.number]).columns)
                numeric_cols = list(source_numeric.intersection(target_numeric))

                if not numeric_cols:
                    # Create worksheet with message if no common numeric columns
                    agg_sheet = writer.book.create_sheet('AggregationCheck')
                    agg_sheet.cell(row=1, column=1, value='No common numeric columns found for aggregation check')
                else:
                    # Calculate aggregations for common numeric columns
                    source_aggs = self._calculate_aggregations(source_df, numeric_cols)
                    target_aggs = self._calculate_aggregations(target_df, numeric_cols)
                    
                    agg_comparison = []
                    for col in numeric_cols:
                        try:
                            source_row = source_aggs[source_aggs['Column'] == col].iloc[0]
                            target_row = target_aggs[target_aggs['Column'] == col].iloc[0]
                            
                            for metric in ['Sum', 'Mean', 'Min', 'Max', 'StdDev']:
                                # Handle NaN values
                                source_val = source_row[metric]
                                target_val = target_row[metric]
                                
                                if pd.isna(source_val) or pd.isna(target_val):
                                    matches = False
                                else:
                                    matches = np.isclose(source_val, target_val, rtol=1e-05)
                                
                                agg_comparison.append({
                                    'Column': col,
                                    'Metric': metric,
                                    'Source': source_val,
                                    'Target': target_val,
                                    'Result': 'PASS' if matches else 'FAIL'
                                })
                        except Exception as e:
                            logger.error(f"Error comparing column {col}: {str(e)}")
                            # Add error entry
                            for metric in ['Sum', 'Mean', 'Min', 'Max', 'StdDev']:
                                agg_comparison.append({
                                    'Column': col,
                                    'Metric': metric,
                                    'Source': 'ERROR',
                                    'Target': 'ERROR',
                                    'Result': 'FAIL'
                                })
                    
                    agg_df = pd.DataFrame(agg_comparison)
                    agg_df.to_excel(writer, sheet_name='AggregationCheck', index=False)
                    
                    # Apply conditional formatting
                    agg_sheet = writer.sheets['AggregationCheck']
                    for idx, result in enumerate(agg_df['Result'], start=2):
                        cell = agg_sheet.cell(row=idx, column=5)
                        self._style_excel_cell(cell, result == 'PASS')

                # 3. Distinct Check Tab
                # Create worksheet
                distinct_sheet = writer.book.create_sheet('DistinctCheck')
                
                # Write headers
                headers = ['Column', 'Source Distinct Count', 'Target Distinct Count', 'Count Match', 'Values Match', 'Source Values', 'Target Values']
                for col_idx, header in enumerate(headers, start=1):
                    cell = distinct_sheet.cell(row=1, column=col_idx, value=header)
                    cell.font = Font(bold=True)
                
                # Get non-numeric columns
                non_numeric_cols = source_df.select_dtypes(exclude=[np.number]).columns
                
                # Process each column
                row_idx = 2
                for col in non_numeric_cols:
                    try:
                        # Get source values
                        source_vals = sorted(source_df[col].fillna('NULL').astype(str).unique())
                        source_count = len(source_vals)
                        source_display = ', '.join(source_vals[:10])
                        if source_count > 10:
                            source_display += '...'
                        
                        # Get target values
                        if col in target_df.columns:
                            target_vals = sorted(target_df[col].fillna('NULL').astype(str).unique())
                            target_count = len(target_vals)
                            target_display = ', '.join(target_vals[:10])
                            if target_count > 10:
                                target_display += '...'
                        else:
                            target_vals = []
                            target_count = 0
                            target_display = 'Column not found'
                        
                        # Compare values
                        count_match = 'PASS' if source_count == target_count else 'FAIL'
                        values_match = 'PASS' if set(source_vals) == set(target_vals) else 'FAIL'
                        
                        # Write row
                        values = [str(col), source_count, target_count, count_match, values_match, source_display, target_display]
                        for col_idx, value in enumerate(values, start=1):
                            cell = distinct_sheet.cell(row=row_idx, column=col_idx, value=value)
                            if col_idx in [4, 5]:  # Format match columns
                                fill_color = PatternFill(start_color='90EE90' if value == 'PASS' else 'FFB6C1', 
                                                       end_color='90EE90' if value == 'PASS' else 'FFB6C1',
                                                       fill_type='solid')
                                cell.fill = fill_color
                        row_idx += 1
                    except Exception as e:
                        logger.error(f"Error processing column {col}: {str(e)}")
                        continue
                
                # Write message if no columns processed
                if row_idx == 2:
                    distinct_sheet.cell(row=2, column=1, value='No non-numeric columns found')
                
                # Adjust column widths
                for col_idx in range(1, len(headers) + 1):
                    distinct_sheet.column_dimensions[get_column_letter(col_idx)].width = 20
                
            logger.info(f"Enhanced regression report generated: {report_path}")
            return str(report_path)
            
        except Exception as e:
            logger.error(f"Error generating regression report: {str(e)}")
            raise

    def generate_difference_report(self, source_df: pd.DataFrame, target_df: pd.DataFrame, 
                                 join_columns: List[str], source_query: str = None, 
                                 target_query: str = None) -> str:
        """Generate enhanced side-by-side difference report."""
        report_path = None
        try:
            if not isinstance(source_df, pd.DataFrame) or not isinstance(target_df, pd.DataFrame):
                raise ValueError("Both source and target must be pandas DataFrames")
                
            # Execute queries if provided
            try:
                if source_query:
                    logger.info(f"Executing source query: {source_query}")
                    source_df = self.execute_query(source_df, source_query)
                    if source_df.empty:
                        logger.warning("Source query returned no results")
                
                if target_query:
                    logger.info(f"Executing target query: {target_query}")
                    target_df = self.execute_query(target_df, target_query)
                    if target_df.empty:
                        logger.warning("Target query returned no results")
            except Exception as e:
                logger.error(f"Error executing queries: {str(e)}")
                raise ValueError(f"Query execution failed: {str(e)}")
                
            if source_df.empty or target_df.empty:
                logger.info("No data to compare in difference report")
                return None
                
            if not join_columns or not isinstance(join_columns, list):
                raise ValueError("join_columns must be a non-empty list")

            logger.info(f"Source columns: {list(source_df.columns)}")
            logger.info(f"Target columns: {list(target_df.columns)}")
            logger.info(f"Join columns requested: {join_columns}")
            
            # Case-insensitive column mapping
            source_cols_lower = {col.lower(): col for col in source_df.columns}
            target_cols_lower = {col.lower(): col for col in target_df.columns}
            
            logger.info(f"Source columns (lowercase): {list(source_cols_lower.keys())}")
            logger.info(f"Target columns (lowercase): {list(target_cols_lower.keys())}")
            
            # Create mapping for join columns
            join_col_mapping = {}
            missing_in_source = []
            missing_in_target = []
            
            try:
                for join_col in join_columns:
                    join_col_lower = join_col.lower()
                    logger.info(f"Processing join column: {join_col} (lowercase: {join_col_lower})")
                    
                    # Find matching source column
                    source_match = source_cols_lower.get(join_col_lower)
                    if not source_match:
                        logger.warning(f"Join column {join_col} not found in source columns")
                        missing_in_source.append(join_col)
                        continue
                    
                    # Find matching target column
                    target_match = target_cols_lower.get(join_col_lower)
                    if not target_match:
                        logger.warning(f"Join column {join_col} not found in target columns")
                        missing_in_target.append(join_col)
                        continue
                    
                    logger.info(f"Mapped {join_col}: source={source_match}, target={target_match}")
                    join_col_mapping[source_match] = target_match
                
                # Check for missing columns after processing all join columns
                if missing_in_source or missing_in_target:
                    error_msg = []
                    if missing_in_source:
                        error_msg.append(f"Join columns missing in source: {', '.join(missing_in_source)}")
                    if missing_in_target:
                        error_msg.append(f"Join columns missing in target: {', '.join(missing_in_target)}")
                    raise ValueError('\n'.join(error_msg))
                    
            except Exception as e:
                logger.error(f"Error during column mapping: {str(e)}")
                raise ValueError(f"Failed to process join columns: {str(e)}")
            
            # Map columns with case-insensitive matching
            logger.info("=== Column Mapping Process ===")
            logger.info(f"Source columns: {list(source_df.columns)}")
            logger.info(f"Target columns: {list(target_df.columns)}")
            logger.info(f"Join columns: {join_columns}")

            # Create case-insensitive column maps
            source_cols = {col.strip().lower(): col for col in source_df.columns}
            target_cols = {col.strip().lower(): col for col in target_df.columns}
            
            # Map columns
            join_col_mapping = {}
            missing_cols = []
            
            for col in join_columns:
                col_lower = col.strip().lower()
                source_match = source_cols.get(col_lower)
                target_match = target_cols.get(col_lower)
                
                if source_match and target_match:
                    join_col_mapping[source_match] = target_match
                    logger.info(f"Mapped: {source_match} -> {target_match}")
                else:
                    if not source_match:
                        missing_cols.append(f"'{col}' (in source)")
                    if not target_match:
                        missing_cols.append(f"'{col}' (in target)")

            if missing_cols:
                error_msg = f"Column mapping failed. Missing columns: {', '.join(missing_cols)}"
                logger.error(error_msg)
                raise ValueError(error_msg)

            if not join_col_mapping:
                error_msg = "No valid column mappings found between source and target"
                logger.error(error_msg)
                raise ValueError(error_msg)

            # Create renamed target dataframe
            target_df_renamed = target_df.copy()
            rename_map = {v: k for k, v in join_col_mapping.items()}
            
            logger.info("=== Column Mapping ===")
            logger.info(f"Final mapping: {join_col_mapping}")
            logger.info(f"Rename map: {rename_map}")

            # Rename target columns
            target_df_renamed = target_df_renamed.rename(columns=rename_map)
            
            # Update join columns to use source column names
            join_columns = list(join_col_mapping.keys())

            logger.info("=== Final State ===")
            logger.info(f"Final join columns: {join_columns}")
            logger.info(f"Renamed target columns: {target_df_renamed.columns.tolist()}")

            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            report_path = self.output_dir / f"DifferenceReport_{timestamp}.xlsx"

            # Constants for Excel limitations - reduced for better performance
            MAX_ROWS = 100000  # Significantly reduced from Excel's limit for stability
            CHUNK_SIZE = 90000  # Slightly less than max for headers and formatting
            
            # Process data in chunks to avoid memory issues
            dfs_to_process = []
            
            # Handle data types for join columns
            for col in join_columns:
                source_type = source_df[col].dtype
                target_type = target_df[col].dtype
                
                # Check if either column contains float values
                is_source_float = source_df[col].apply(lambda x: isinstance(x, float) or (isinstance(x, str) and '.' in x)).any()
                is_target_float = target_df[col].apply(lambda x: isinstance(x, float) or (isinstance(x, str) and '.' in x)).any()
                
                if is_source_float or is_target_float:
                    # Convert both columns to float
                    logger.info(f"Converting column {col} to float type")
                    try:
                        source_df[col] = pd.to_numeric(source_df[col], errors='coerce')
                        target_df[col] = pd.to_numeric(target_df[col], errors='coerce')
                    except Exception as e:
                        logger.warning(f"Could not convert column {col} to float: {str(e)}")
                elif source_type != target_type:
                    # For non-float columns, convert to string if types don't match
                    logger.info(f"Converting column {col} to string type")
                    source_df[col] = source_df[col].astype(str)
                    target_df[col] = target_df[col].astype(str)

            for start_idx in range(0, len(source_df), CHUNK_SIZE):
                try:
                    # Get chunks of both dataframes
                    source_chunk = source_df.iloc[start_idx:start_idx + CHUNK_SIZE]
                    
                    try:
                        logger.info(f"Processing chunk {chunk_idx + 1}")
                        
                        # Verify columns before merge
                        for col in join_columns:
                            if col not in source_chunk.columns:
                                raise ValueError(f"Join column '{col}' missing from source chunk")
                            if col not in target_df_renamed.columns:
                                raise ValueError(f"Join column '{col}' missing from target")
                            
                            # Convert to string if data types don't match
                            if source_chunk[col].dtype != target_df_renamed[col].dtype:
                                logger.info(f"Converting {col} to string type for comparison")
                                source_chunk[col] = source_chunk[col].astype(str)
                                target_df_renamed[col] = target_df_renamed[col].astype(str)
                        
                        # Perform merge with explicit column list
                        merge_cols = join_columns.copy()  # Use only mapped columns
                        chunk_merged = source_chunk.merge(
                            target_df_renamed,
                            on=merge_cols,
                            how='outer',
                            indicator=True,
                            suffixes=('_source', '_target')
                        )
                        
                        logger.info(f"Merge successful - Shape: {chunk_merged.shape}")
                        
                    except Exception as merge_error:
                        logger.error(f"Error during merge operation: {str(merge_error)}")
                        raise ValueError(f"Failed to merge data: {str(merge_error)}")
                    
                    # Create comparison status column
                    chunk_merged['Status'] = chunk_merged['_merge'].map({
                        'left_only': 'Deleted',
                        'right_only': 'Inserted',
                        'both': 'Updated'
                    })
                    
                    # Remove the merge indicator column
                    chunk_merged = chunk_merged.drop('_merge', axis=1)
                    
                    # Only keep rows with differences
                    diff_rows = chunk_merged[chunk_merged['Status'] != 'Updated']
                    if not diff_rows.empty:
                        dfs_to_process.append(diff_rows)
                        
                except Exception as chunk_error:
                    logger.error(f"Error processing chunk starting at index {start_idx}: {str(chunk_error)}")
                    continue
            
            try:
                # If there are no differences
                if not dfs_to_process:
                    logger.info("No differences found between source and target")
                    with open(report_path, 'w') as f:
                        f.write("No differences found between source and target datasets.")
                    return str(report_path)
                
                # Combine all difference chunks
                merged = pd.concat(dfs_to_process, ignore_index=True)
                
                # Define status colors
                status_colors = {
                    'Deleted': 'FFB6C1',     # Light pink
                    'Left Only': 'FFB6C1',   # Light pink
                    'Inserted': '90EE90',    # Light green
                    'Right Only': '90EE90',  # Light green
                    'Updated': 'FFD700'      # Gold
                }
                
                # Calculate number of chunks needed
                total_rows = len(merged)
                num_chunks = (total_rows - 1) // CHUNK_SIZE + 1
                
            except Exception as e:
                logger.error(f"Error processing difference chunks: {str(e)}")
                raise ValueError(f"Failed to process difference data: {str(e)}")
            
            # Save to Excel with xlsxwriter engine for better compatibility
            with pd.ExcelWriter(str(report_path), engine='xlsxwriter') as writer:
                for chunk_idx in range(num_chunks):
                    try:
                        start_idx = chunk_idx * CHUNK_SIZE
                        end_idx = min((chunk_idx + 1) * CHUNK_SIZE, total_rows)
                        
                        # Get chunk of data
                        chunk = merged.iloc[start_idx:end_idx]
                        
                        # Create sheet name
                        sheet_name = 'Differences' if num_chunks == 1 else f'Differences_{chunk_idx + 1}'
                        
                        # Clean up column names to avoid Excel errors
                        chunk.columns = [str(col).strip().replace('/', '_').replace('\\', '_') for col in chunk.columns]
                        
                        # Write chunk to Excel
                        chunk.to_excel(writer, sheet_name=sheet_name, index=False)
                        
                        # Get worksheet and create formats
                        worksheet = writer.sheets[sheet_name]
                        workbook = writer.book
                        
                        # Create formats for different statuses
                        formats = {
                            status: workbook.add_format({
                                'bg_color': color,
                                'border': 1
                            }) for status, color in status_colors.items()
                        }
                        
                        # Find Status column index
                        status_col_idx = None
                        for idx, col in enumerate(chunk.columns):
                            if col == 'Status':
                                status_col_idx = idx
                                break
                        
                        if status_col_idx is not None:
                            # Apply conditional formatting based on status
                            for row_idx, status in enumerate(chunk['Status'], start=1):
                                worksheet.write(row_idx, status_col_idx, status, formats.get(status))
                        
                        # Add summary at top of each sheet
                        summary_data = chunk['Status'].value_counts().to_dict()
                        summary_col = len(chunk.columns) + 2
                        
                        # Create header format
                        header_format = workbook.add_format({
                            'bold': True,
                            'font_size': 11,
                            'border': 1
                        })
                        
                        # Write summary
                        worksheet.write(0, summary_col, "Summary:", header_format)
                        for i, (status, count) in enumerate(summary_data.items()):
                            worksheet.write(i + 1, summary_col, f"{status}: {count}", formats.get(status))
                            
                        # Auto-adjust column widths
                        for col_num, value in enumerate(chunk.columns.values):
                            max_length = max(
                                chunk[value].astype(str).str.len().max(),
                                len(str(value))
                            )
                            worksheet.set_column(col_num, col_num, min(max_length + 2, 50))
                            
                    except Exception as sheet_error:
                        logger.error(f"Error writing sheet {sheet_name}: {str(sheet_error)}")
                        continue
                
                try:
                    # Create summary sheet
                    summary_sheet = writer.book.add_worksheet('Summary')
                    
                    # Create formats
                    header_format = writer.book.add_format({
                        'bold': True,
                        'font_size': 11,
                        'border': 1,
                        'bg_color': '#D3D3D3'
                    })
                    
                    cell_format = writer.book.add_format({
                        'border': 1
                    })
                    
                    # Write headers
                    summary_sheet.write(0, 0, 'Description', header_format)
                    summary_sheet.write(0, 1, 'Value', header_format)
                    
                    # Write summary data
                    summary_data = [
                        ('Total Rows', total_rows),
                        ('Number of Sheets', num_chunks),
                        ('Rows per Sheet', CHUNK_SIZE),
                        ('Deleted Records', sum(merged['Status'] == 'Deleted')),
                        ('Inserted Records', sum(merged['Status'] == 'Inserted')),
                        ('Updated Records', sum(merged['Status'] == 'Updated'))
                    ]
                    
                    for row_idx, (desc, value) in enumerate(summary_data, start=1):
                        summary_sheet.write(row_idx, 0, desc, cell_format)
                        summary_sheet.write(row_idx, 1, value, cell_format)
                    
                    # Auto-adjust column widths
                    summary_sheet.set_column(0, 0, 20)  # Description column
                    summary_sheet.set_column(1, 1, 15)  # Value column
                        
                except Exception as summary_error:
                    logger.error(f"Error creating summary sheet: {str(summary_error)}")
            
            logger.info(f"Difference report generated: {report_path}")
            return str(report_path)
            
        except Exception as e:
            logger.error(f"Error generating difference report: {str(e)}")
            raise

    def generate_datacompy_report(self, comparison_results: Dict[str, Any]) -> str:
        """Generate detailed DataCompy report with proper formatting."""
        try:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            report_path = self.output_dir / f"DataCompy_{timestamp}.txt"
            
            with open(report_path, 'w', encoding='utf-8') as f:
                # Write header with timestamp
                f.write("=" * 80 + "\n")
                f.write("DataCompy Comparison Report\n")
                f.write(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
                f.write("=" * 80 + "\n\n")

                # Write summary section
                f.write("SUMMARY\n")
                f.write("-" * 80 + "\n")
                f.write(f"Source Rows: {comparison_results.get('row_counts', {}).get('source_count', 'N/A')}\n")
                f.write(f"Target Rows: {comparison_results.get('row_counts', {}).get('target_count', 'N/A')}\n")
                
                # Write column details
                f.write("\nCOLUMN ANALYSIS\n")
                f.write("-" * 80 + "\n")
                for col, details in comparison_results.get('column_summary', {}).items():
                    f.write(f"\nColumn: {col}\n")
                    f.write("  Source:\n")
                    f.write(f"    Null Count: {details.get('source_null_count', 'N/A')}\n")
                    f.write(f"    Unique Values: {details.get('source_unique_count', 'N/A')}\n")
                    if 'source_mean' in details:
                        f.write(f"    Mean: {details['source_mean']}\n")
                        f.write(f"    Sum: {details.get('source_sum', 'N/A')}\n")
                    
                    f.write("  Target:\n")
                    f.write(f"    Null Count: {details.get('target_null_count', 'N/A')}\n")
                    f.write(f"    Unique Values: {details.get('target_unique_count', 'N/A')}\n")
                    if 'target_mean' in details:
                        f.write(f"    Mean: {details['target_mean']}\n")
                        f.write(f"    Sum: {details.get('target_sum', 'N/A')}\n")
                    f.write("\n")
                
                # Write unmatched rows summary
                f.write("\nUNMATCHED ROWS SUMMARY\n")
                f.write("-" * 80 + "\n")
                
                source_unmatched = comparison_results.get('source_unmatched_rows', pd.DataFrame())
                target_unmatched = comparison_results.get('target_unmatched_rows', pd.DataFrame())
                
                f.write(f"Source Unmatched Count: {len(source_unmatched)}\n")
                f.write(f"Target Unmatched Count: {len(target_unmatched)}\n\n")

                if not source_unmatched.empty:
                    f.write("Sample of Source Unmatched Rows (first 5):\n")
                    f.write("-" * 40 + "\n")
                    sample = source_unmatched.head(5).to_string()
                    f.write(sample + "\n\n")

                if not target_unmatched.empty:
                    f.write("Sample of Target Unmatched Rows (first 5):\n")
                    f.write("-" * 40 + "\n")
                    sample = target_unmatched.head(5).to_string()
                    f.write(sample + "\n\n")

                # Write match status summary
                f.write("\nMATCH STATUS SUMMARY\n")
                f.write("-" * 80 + "\n")
                f.write(f"Rows Match: {'Yes' if comparison_results.get('rows_match', False) else 'No'}\n")
                f.write(f"Columns Match: {'Yes' if comparison_results.get('columns_match', False) else 'No'}\n")
                f.write(f"Data Matches: {'Yes' if comparison_results.get('match_status', False) else 'No'}\n")

                # Write any additional details from datacompy
                if 'datacompy_report' in comparison_results:
                    f.write("\nDETAILED COMPARISON\n")
                    f.write("-" * 80 + "\n")
                    f.write(comparison_results['datacompy_report'])
            
            logger.info(f"DataCompy report generated: {report_path}")
            return str(report_path)
            
        except Exception as e:
            logger.error(f"Error generating DataCompy report: {str(e)}")
            raise

    def generate_ydata_report(self, comparison_results: Dict[str, Any]) -> str:
        """Generate Y-Data Profiling comparison report."""
        try:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            report_path = self.output_dir / f"Comparison_profile_{timestamp}.html"
            
            # Create resources directory
            resources_dir = self.output_dir / 'resources'
            resources_dir.mkdir(exist_ok=True)
            
            # Get profile data
            source_profile = comparison_results.get('source_profile', {})
            target_profile = comparison_results.get('target_profile', {})
            
            # Generate comparison HTML
            with open(report_path, 'w', encoding='utf-8') as f:
                html_content = '''
                <!DOCTYPE html>
                <html>
                <head>
                    <meta charset="utf-8">
                    <title>Data Profile Comparison Report</title>
                    <style>
                        body { font-family: Arial, sans-serif; margin: 2em; line-height: 1.6; }
                        h1, h2 { color: #333; }
                        .container { max-width: 1200px; margin: 0 auto; }
                        .section { margin-bottom: 2em; padding: 1em; border: 1px solid #ddd; border-radius: 4px; }
                        .diff { background-color: #fff3cd; padding: 0.5em; }
                        table { width: 100%; border-collapse: collapse; margin: 1em 0; }
                        th, td { padding: 8px; text-align: left; border-bottom: 1px solid #ddd; }
                        th { background-color: #f5f5f5; }
                        .highlight { background-color: #ffe3e3; }
                    </style>
                </head>
                <body>
                    <div class="container">
                        <h1>Data Profile Comparison Report</h1>
                '''
                
                # Add timestamp
                html_content += f'<p>Generated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}</p>'
                
                # Basic Statistics Section
                html_content += '''
                    <div class="section">
                        <h2>Basic Statistics</h2>
                        <table>
                            <tr>
                                <th>Metric</th>
                                <th>Source</th>
                                <th>Target</th>
                                <th>Difference</th>
                            </tr>
                '''
                
                # Add basic statistics comparison
                metrics = ['row_count', 'column_count', 'duplicate_rows', 'missing_cells']
                for metric in metrics:
                    source_val = source_profile.get(metric, 'N/A')
                    target_val = target_profile.get(metric, 'N/A')
                    diff = ''
                    if isinstance(source_val, (int, float)) and isinstance(target_val, (int, float)):
                        diff = target_val - source_val
                    
                    html_content += f'''
                        <tr>
                            <td>{metric.replace('_', ' ').title()}</td>
                            <td>{source_val}</td>
                            <td>{target_val}</td>
                            <td>{diff if diff != '' else 'N/A'}</td>
                        </tr>
                    '''
                
                html_content += '</table></div>'
                
                # Column Analysis Section
                html_content += '''
                    <div class="section">
                        <h2>Column Analysis</h2>
                '''
                
                # Get all columns from both profiles
                all_columns = set(source_profile.get('columns', {}).keys()) | set(target_profile.get('columns', {}).keys())
                
                for column in sorted(all_columns):
                    source_col = source_profile.get('columns', {}).get(column, {})
                    target_col = target_profile.get('columns', {}).get(column, {})
                    
                    html_content += f'''
                        <div class="section">
                            <h3>Column: {column}</h3>
                            <table>
                                <tr>
                                    <th>Metric</th>
                                    <th>Source</th>
                                    <th>Target</th>
                                </tr>
                    '''
                    
                    # Compare column metrics
                    col_metrics = ['type', 'unique_count', 'missing_count', 'min', 'max', 'mean', 'std']
                    for metric in col_metrics:
                        source_val = source_col.get(metric, 'N/A')
                        target_val = target_col.get(metric, 'N/A')
                        highlight = ' class="highlight"' if source_val != target_val else ''
                        
                        html_content += f'''
                            <tr{highlight}>
                                <td>{metric.replace('_', ' ').title()}</td>
                                <td>{source_val}</td>
                                <td>{target_val}</td>
                            </tr>
                        '''
                    
                    html_content += '</table></div>'
                
                # Close main container and body
                html_content += '''
                    </div>
                </body>
                </html>
                '''
                
                f.write(html_content)
            
            logger.info(f"Y-Data Profile report generated: {report_path}")
            return str(report_path)
            
        except Exception as e:
            logger.error(f"Error generating Y-Data Profile report: {str(e)}")
            raise

    def create_report_archive(self, report_paths: Dict[str, str]) -> str:
        """Create a ZIP archive containing all reports."""
        try:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            zip_path = self.output_dir / f"all_reports_{timestamp}.zip"
            
            with zipfile.ZipFile(str(zip_path), 'w') as zipf:
                for report_type, path in report_paths.items():
                    if path and os.path.exists(path):
                        # Add file to zip with its original name
                        zipf.write(path, os.path.basename(path))
                        
                        # For HTML reports, also add any associated resources
                        if path.endswith('.html'):
                            resources_dir = Path(path).parent / 'resources'
                            if resources_dir.exists():
                                for resource in resources_dir.rglob('*'):
                                    if resource.is_file():
                                        zipf.write(
                                            resource,
                                            os.path.join('resources', resource.relative_to(resources_dir))
                                        )
            
            logger.info(f"Report archive created: {zip_path}")
            return str(zip_path)
            
        except Exception as e:
            logger.error(f"Error creating report archive: {str(e)}")
            raise
